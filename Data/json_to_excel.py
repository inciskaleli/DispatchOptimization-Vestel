# -*- coding: utf-8 -*-
"""
Creates a single Excel file with:
  - Assignments (per-tech rows) + dist_from_prev_m, Nonavailabilities, Technicians, Meta
  - Suggestions_* (if present)
  - Tech_Load_By_Slot (counts & minutes)
  - appt_eligibility_by_slot
  - Tech_Total_Distance  <-- NEW (sum of per-tech hop distances)

'eligible_with_free_space' = for each appointment and slot on that day:
  number of ELIGIBLE technicians who would have at least the appointment's
  duration minutes available in that slot IF THIS APPOINTMENT WERE NOT ASSIGNED.
"""

import json
import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Tuple, Optional

import pandas as pd
from openpyxl.styles import Font
import math
from typing import Optional
 

sayi = 14
import argparse

parser = argparse.ArgumentParser()
parser.add_argument("--input_result", required=True, help="Path to result JSON")
parser.add_argument("--input_dataloader", required=True, help="Path to dataloader JSON")
parser.add_argument("--output_xlsx", required=False, help="Path to output XLSX (optional)")
args = parser.parse_args()

INPUT_RESULT_JSON = args.input_result
INPUT_DATALOADER  = args.input_dataloader
OUTPUT_XLSX       = args.output_xlsx or Path(INPUT_RESULT_JSON).with_suffix(".xlsx")

SLOT_CAPACITY_MIN  = 120  # 2 hours per slot
# ------------------------


SLOT_DEFS: List[Tuple[str, int, int]] = [
    ("08:00-10:00", 8, 10),
    ("10:00-12:00", 10, 12),
    ("13:00-15:00", 13, 15),
    ("15:00-17:00", 15, 17),
    ("17:00-19:00", 17, 19),
    ("19:00-21:00", 19, 21),
    ("21:00-23:00", 21, 23),
]
SLOT_LABELS = [s[0] for s in SLOT_DEFS]


# ----------------- Common helpers -----------------
def _to_list(v):
    return v if isinstance(v, list) else ([] if v is None else [v])


def parse_iso_z(s: str) -> datetime:
    """Parse '...Z' or '...+00:00' → aware UTC datetime."""
    s = (s or "").strip().replace(".000Z", "Z")
    if s.endswith("Z"):
        s = s[:-1] + "+00:00"
    return datetime.fromisoformat(s)


def slot_bounds_for_date(d: datetime) -> list[tuple[datetime, datetime, str]]:
    base = datetime(d.year, d.month, d.day, tzinfo=timezone.utc)
    out = []
    for label, h1, h2 in SLOT_DEFS:
        sdt = base.replace(hour=h1, minute=0, second=0, microsecond=0)
        edt = base.replace(hour=h2, minute=0, second=0, microsecond=0)
        out.append((sdt, edt, label))
    return out


def label_for_window(start_iso: str, end_iso: str) -> Optional[str]:
    """Return slot label if the arrival window exactly matches a standard slot."""
    try:
        s = parse_iso_z(start_iso)
        e = parse_iso_z(end_iso)
    except Exception:
        return None
    for sdt, edt, label in slot_bounds_for_date(s):
        if s == sdt and e == edt:
            return label
    return None


def overlap_minutes(a_start: datetime, a_end: datetime, b_start: datetime, b_end: datetime) -> int:
    latest_start = max(a_start, b_start)
    earliest_end = min(a_end, b_end)
    delta = (earliest_end - latest_start).total_seconds() / 60.0
    return max(0, int(round(delta)))


# ----------------- Matrix helpers (NEW) -----------------
def _norm_coord_str(coord: Optional[str]) -> Optional[str]:
    """Normalize 'lon,lat'/'lon;lat' → 'lon,lat' trimmed; return None if invalid/missing."""
    if not isinstance(coord, str):
        return None
    s = coord.strip().replace(";", ",")
    if "," not in s:
        return None
    return s


def _reformat_coord_str(coord: Optional[str], nd: int = 6) -> Optional[str]:
    """Try to parse and reformat to fixed decimals to improve matrix key matching."""
    s = _norm_coord_str(coord)
    if s is None:
        return None
    try:
        a, b = [p.strip() for p in s.split(",", 1)]
        lon = float(a)
        lat = float(b)
        return f"{lon:.{nd}f},{lat:.{nd}f}"
    except Exception:
        return s  # fall back to original normalized


def _matrix_distance_m(matrix_distance: Dict[str, Dict[str, Any]],
                       origin_coord: Optional[str],
                       dest_coord: Optional[str]) -> int:
    """
    Lookup distance in meters from matrix_distance[origin][dest].
    Tries exact normalized key, then a 6-decimal reformatted key.
    Returns 0 if any key missing.
    """
    if not matrix_distance:
        return 0
    o1 = _norm_coord_str(origin_coord)
    d1 = _norm_coord_str(dest_coord)
    if not o1 or not d1:
        return 0

    # direct
    row = matrix_distance.get(o1)
    if isinstance(row, dict) and d1 in row:
        try:
            return int(row[d1])
        except Exception:
            return 0

    # try reformatted keys (common when sources have different float precision)
    o2 = _reformat_coord_str(o1)
    d2 = _reformat_coord_str(d1)
    row2 = matrix_distance.get(o2) if o2 else None
    if isinstance(row2, dict) and d2 in row2:
        try:
            return int(row2[d2])
        except Exception:
            return 0

    # last attempt: from original row to reformatted dest (or vice versa)
    if isinstance(row, dict) and d2 and d2 in row:
        try:
            return int(row[d2])
        except Exception:
            return 0
    if o2 and isinstance(matrix_distance.get(o2), dict) and d1 in matrix_distance[o2]:
        try:
            return int(matrix_distance[o2][d1])
        except Exception:
            return 0

    # not found
    return 0


# ----------------- Flatten result JSON -----------------
def flatten_assignments(assignments: List[Dict[str, Any]]) -> pd.DataFrame:
    rows = []
    for a in assignments or []:
        route = a.get("route") or {}
        tech_ids = _to_list(a.get("technician_ids"))
        rows.append({
            "appointment_id": str(a.get("id")),
            "status": a.get("status"),
            "start": a.get("start"),
            "end": a.get("end"),
            "technician_ids": tech_ids,
            "route_distance": route.get("distance"),
            "route_duration": route.get("duration"),
        })
    return pd.DataFrame(rows)


def flatten_outliers(result_obj: Dict[str, Any]) -> pd.DataFrame:
    """
    Collects explicitly unassigned tasks if they come under 'outliers' or 'unassigned'.
    Falls back to empty if those keys are absent (in many cases outliers are already
    inside 'assignments' with status='outlier').
    """
    rows = []
    buckets = []
    if isinstance(result_obj, dict):
        for key in ("outliers", "unassigned"):
            if isinstance(result_obj.get(key), list):
                buckets.append(result_obj.get(key))
    for bucket in buckets:
        for o in bucket or []:
            route = o.get("route") or {}
            rows.append({
                "appointment_id": str(o.get("id")),
                "status": (o.get("status") or "outlier"),
                "start": o.get("start"),
                "end": o.get("end"),
                "technician_ids": _to_list(o.get("technician_ids")) if o.get("technician_ids") else [],
                "route_distance": route.get("distance"),
                "route_duration": route.get("duration"),
            })
    return pd.DataFrame(rows)


def flatten_nonavail(nonavail: List[Dict[str, Any]]) -> pd.DataFrame:
    rows = []
    for n in nonavail or []:
        nas = n.get("non_availabilities") or []
        if not nas:
            rows.append({"technician_id": n.get("technician_id"),
                         "na_start": None, "na_end": None, "reason": None})
        for item in nas:
            rows.append({
                "technician_id": n.get("technician_id"),
                "na_start": item.get("start"),
                "na_end": item.get("end"),
                "reason": item.get("reason") if isinstance(item, dict) else None,
            })
    return pd.DataFrame(rows)


def flatten_technicians(techs: List[Dict[str, Any]]) -> pd.DataFrame:
    rows = []
    for t in techs or []:
        lb = t.get("lunch_break") or {}
        rows.append({
            "technician_id": t.get("id"),
            "lunch_start": lb.get("start"),
            "lunch_end": lb.get("end"),
        })
    return pd.DataFrame(rows)


def flatten_suggestions(sugg: Dict[str, Any]) -> Dict[str, pd.DataFrame]:
    dfs = {}
    if not sugg:
        return dfs
    overtime = sugg.get("overtime") or []
    rows_ot = []
    for o in overtime:
        appts = _to_list(o.get("appointment_ids"))
        rows_ot.append({
            "technician_id": o.get("technician_id"),
            "appointment_ids": ",".join(appts)
        })
    if rows_ot:
        dfs["Suggestions_Overtime"] = pd.DataFrame(rows_ot)
    buffer_slots = sugg.get("buffer_slots") or []
    rows_bs = []
    for b in buffer_slots:
        rows_bs.append(b if isinstance(b, dict) else {"value": b})
    if rows_bs:
        dfs["Suggestions_BufferSlots"] = pd.DataFrame(rows_bs)
    return dfs


# --------- Tech load by slot (counts & minutes) ----------
def _slot_label_for_hour(h: int) -> Optional[str]:
    for label, start_h, end_h in SLOT_DEFS:
        if start_h <= h < end_h:
            return label
    return None


def _prep_slot_base(assignments_df: pd.DataFrame) -> pd.DataFrame:
    df = assignments_df.copy()
    if df.empty:
        return df
    df = df.explode("technician_ids").rename(columns={"technician_ids": "tech_id"})
    df = df[df["tech_id"].notna() & (df["tech_id"].astype(str) != "")]
    starts = pd.to_datetime(df["start"], errors="coerce", utc=True)
    ends   = pd.to_datetime(df["end"], errors="coerce", utc=True)
    df["start_dt"] = starts
    df["end_dt"]   = ends
    df["duration_min"] = ((ends - starts).dt.total_seconds() / 60.0).fillna(0).clip(lower=0)
    df["slot"] = starts.dt.hour.apply(lambda h: _slot_label_for_hour(int(h)) if pd.notna(h) else None)
    return df[df["slot"].isin(SLOT_LABELS)]


def build_tech_slot_counts(assignments_df: pd.DataFrame) -> pd.DataFrame:
    if assignments_df.empty:
        return pd.DataFrame(columns=["Tech Id"] + SLOT_LABELS)
    df = _prep_slot_base(assignments_df)
    pivot = (df.groupby(["tech_id", "slot"])
               .size()
               .unstack(fill_value=0)
               .reindex(columns=SLOT_LABELS, fill_value=0)
               .reset_index()
               .rename(columns={"tech_id": "Tech Id"}))
    for col in SLOT_LABELS:
        if col not in pivot.columns:
            pivot[col] = 0
    return pivot[["Tech Id"] + SLOT_LABELS]


def build_tech_slot_minutes(assignments_df: pd.DataFrame) -> pd.DataFrame:
    if assignments_df.empty:
        return pd.DataFrame(columns=["Tech Id"] + SLOT_LABELS)
    df = _prep_slot_base(assignments_df)
    pivot = (df.groupby(["tech_id", "slot"])["duration_min"]
               .sum()
               .unstack(fill_value=0.0)
               .reindex(columns=SLOT_LABELS, fill_value=0.0)
               .reset_index()
               .rename(columns={"tech_id": "Tech Id"}))
    for col in SLOT_LABELS:
        pivot[col] = pivot.get(col, 0).round(0).astype(int)
    return pivot[["Tech Id"] + SLOT_LABELS]


# ------------- Build per-tech-per-date-per-slot used minutes -----------------
def tech_slot_used_minutes_map(assignments_df: pd.DataFrame) -> Dict[str, Dict[Tuple[str, str], int]]:
    """
    Return {tech_id: {(date_str, slot_label): minutes_used}}
    Minutes are computed by precise overlap with the slot on that date.
    """
    out: Dict[str, Dict[Tuple[str, str], int]] = {}
    if assignments_df.empty:
        return out

    df = assignments_df.copy()
    df = df.explode("technician_ids").rename(columns={"technician_ids": "tech_id"})
    df = df[df["tech_id"].notna() & (df["tech_id"].astype(str) != "")]
    df["start_dt"] = pd.to_datetime(df["start"], errors="coerce", utc=True)
    df["end_dt"]   = pd.to_datetime(df["end"], errors="coerce", utc=True)

    for _, row in df.iterrows():
        tech = str(row["tech_id"])
        s = row["start_dt"]; e = row["end_dt"]
        if pd.isna(s) or pd.isna(e):
            continue
        date_key = s.date().isoformat()
        for sdt, edt, label in slot_bounds_for_date(s):
            mins = overlap_minutes(s, e, sdt, edt)
            if mins > 0:
                out.setdefault(tech, {})
                key = (date_key, label)
                out[tech][key] = out[tech].get(key, 0) + mins
    return out


def appt_assigned_overlap_by_tech(assignments_df: pd.DataFrame) -> Dict[Tuple[str, str, str, str], int]:
    """
    (appt_id, tech_id, date_str, slot_label) -> minutes of THIS appt that fall in that slot for that tech.
    Used to "remove" the appt when simulating not-assigned.
    """
    out: Dict[Tuple[str, str, str, str], int] = {}
    if assignments_df.empty:
        return out

    df = assignments_df.copy()
    df = df.explode("technician_ids").rename(columns={"technician_ids": "tech_id"})
    df = df[df["tech_id"].notna() & (df["tech_id"].astype(str) != "")]
    df["start_dt"] = pd.to_datetime(df["start"], errors="coerce", utc=True)
    df["end_dt"]   = pd.to_datetime(df["end"], errors="coerce", utc=True)

    for _, row in df.iterrows():
        apid = str(row.get("appointment_id"))
        tech = str(row["tech_id"])
        s = row["start_dt"]; e = row["end_dt"]
        if pd.isna(s) or pd.isna(e):
            continue
        date_key = s.date().isoformat()
        for sdt, edt, label in slot_bounds_for_date(s):
            mins = overlap_minutes(s, e, sdt, edt)
            if mins > 0:
                out[(apid, tech, date_key, label)] = mins
    return out


# ------------- appt_eligibility_by_slot sheet -----------------
def build_appt_eligibility_by_slot(appts: List[dict],
                                   tech_used_map: Dict[str, Dict[Tuple[str, str], int]],
                                   appt_overlap_map: Dict[Tuple[str, str, str, str], int]) -> pd.DataFrame:
    rows = []
    # precompute appointment durations & arrival labels
    appt_duration: Dict[str, int] = {}
    appt_arrival_label: Dict[str, Optional[str]] = {}
    appt_date_key: Dict[str, str] = {}

    for ap in appts:
        apid = str(ap.get("id", ""))
        try:
            js = parse_iso_z(ap.get("start"))
            je = parse_iso_z(ap.get("end"))
            dur = max(0, int(round((je - js).total_seconds() / 60.0)))
        except Exception:
            dur = 0
        appt_duration[apid] = dur
        aw = ap.get("arrival_window") or {}
        appt_arrival_label[apid] = label_for_window(aw.get("start", ""), aw.get("end", "")) if aw else None
        try:
            aw_s = parse_iso_z(aw.get("start", "")) if aw else parse_iso_z(ap.get("start"))
            appt_date_key[apid] = aw_s.date().isoformat()
        except Exception:
            appt_date_key[apid] = None

    for ap in appts:
        apid = str(ap.get("id", ""))
        bu   = ap.get("business_unit_id") or ""
        elig = [str(x.get("id")) for x in (ap.get("eligible_technicians") or []) if x.get("id") is not None]
        total_eligible = len(elig)
        dur_t = appt_duration.get(apid, 0)
        date_key = appt_date_key.get(apid)
        if not date_key:
            continue

        # find the UTC date to iterate slot bounds
        try:
            aw_s = parse_iso_z((ap.get("arrival_window") or {}).get("start", "")) if ap.get("arrival_window") else parse_iso_z(ap.get("start"))
        except Exception:
            continue

        arrival_label = appt_arrival_label.get(apid)

        for sdt, edt, label in slot_bounds_for_date(aw_s):
            free_count = 0
            for tid in elig:
                used = tech_used_map.get(tid, {}).get((date_key, label), 0)
                # subtract this appointment's own overlap minutes for this tech in that slot (simulate "not assigned")
                used_minus_this = used - appt_overlap_map.get((apid, tid, date_key, label), 0)
                # available minutes
                avail = SLOT_CAPACITY_MIN - max(0, used_minus_this)
                if avail >= dur_t and dur_t > 0:
                    free_count += 1

            rows.append({
                "slot": label,
                "appointment_no": apid,
                "business_unit_id": bu,
                "total_eligible_technicians": total_eligible,
                "eligible_with_free_space": free_count,
                "occupancy_rate": f"{(total_eligible - free_count) / total_eligible:.1f}" if total_eligible > 0 else None,
                "is_arrival_window_row": (label == arrival_label),
            })

    df = pd.DataFrame(rows, columns=[
        "slot",
        "appointment_no",
        "business_unit_id",
        "total_eligible_technicians",
        "eligible_with_free_space",
        "occupancy_rate",
        "is_arrival_window_row",
    ])
    return df


# ----------------- NEW: Build Assignments per-tech with hop distances (from MATRIX) -----------------
def build_assignments_with_dist(assignments_df: pd.DataFrame, dl_obj: Dict[str, Any]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Returns:
      - per-tech assignments dataframe with 'dist_from_prev_m' (from matrix.distance)
      - per-tech total distance dataframe
    Includes UNASSIGNED (status='outlier' or empty technician_ids) rows in the output sheet,
    with blank tech_id and no hop distance.
    Also adds 4 extra columns per appointment:
      'Çağrı tipi', 'Ürün Grubu', 'Yetkinlik Grubu', 'Randevu Tarih Saat'
    """
    if assignments_df.empty:
        empty = assignments_df.copy()
        # add placeholders + new columns so headers are stable
        for c in ["cumulative distance", "total distance", "Çağrı tipi", "Ürün Grubu", "Yetkinlik Grubu", "Randevu Tarih Saat"]:
            empty[c] = None
        cols_final = [
            "appointment_id","status","start","end",
            "tech_id","route_distance","route_duration","dist_from_prev_m",
            "cumulative distance","total distance",
            "Çağrı tipi","Ürün Grubu","Yetkinlik Grubu","Randevu Tarih Saat",
        ]
        return empty.reindex(columns=cols_final), pd.DataFrame(columns=["tech_id", "total_distance_m"])

    # ---------- Build metadata map from dataloader ----------
    def _fmt_aw(aw: Optional[dict]) -> Optional[str]:
        if not isinstance(aw, dict):
            return None
        try:
            sdt = parse_iso_z(aw.get("start", ""))
            edt = parse_iso_z(aw.get("end", ""))
            # e.g., '10.03.2025 15:00:00-17:00:00'
            return f"{sdt.strftime('%d.%m.%Y %H:%M:%S')}-{edt.strftime('%H:%M:%S')}"
        except Exception:
            return None

    appt_meta: Dict[str, Tuple[Optional[str], Optional[str], Optional[str], Optional[str]]] = {}
    for ap in dl_obj.get("appointments", []):
        apid = str(ap.get("id"))
        bu = (ap.get("business_unit_id") or "")
        parts = bu.split("|")
        ct = parts[0] if len(parts) > 0 else None
        pg = parts[1] if len(parts) > 1 else None
        comp = parts[2] if len(parts) > 2 else None
        aw_str = _fmt_aw(ap.get("arrival_window"))
        appt_meta[apid] = (ct, pg, comp, aw_str)

    # ---------- Mark outliers/unassigned on raw frame ----------
    raw = assignments_df.copy()
    status_col = raw.get("status")
    techs_col = raw.get("technician_ids")
    mask_outlier = (status_col.astype(str).str.lower() == "outlier") | techs_col.apply(lambda v: not bool(v))

    # ---------- Matrix & coords ----------
    appt_coord_map: Dict[str, Optional[str]] = {}
    for ap in dl_obj.get("appointments", []):
        appt_coord_map[str(ap.get("id"))] = _norm_coord_str((ap.get("location") or {}).get("coordinate"))

    options = dl_obj.get("options", {}) or {}
    start_at_office = bool(options.get("start_day_at_office", False))
    office_coord_str = _norm_coord_str((options.get("office") or {}).get("coordinate"))

    tech_home_map: Dict[str, Optional[str]] = {}
    for t in dl_obj.get("technicians", []):
        tech_home_map[str(t.get("id"))] = _norm_coord_str((t.get("home") or {}).get("coordinate"))

    matrix_distance: Dict[str, Dict[str, Any]] = (dl_obj.get("matrix") or {}).get("distance") or {}

    # ---------- Assigned rows: explode & compute distances ----------
    assigned_raw = raw.loc[~mask_outlier].copy()
    df = assigned_raw.explode("technician_ids").rename(columns={"technician_ids": "tech_id"})
    df = df[df["tech_id"].notna() & (df["tech_id"].astype(str) != "")]
    if df.empty:
        assigned_out = pd.DataFrame(columns=[
            "appointment_id","status","start","end","tech_id","route_distance","route_duration","dist_from_prev_m"
        ])
    else:
        df["start_dt"] = pd.to_datetime(df["start"], errors="coerce", utc=True)
        df["date"] = df["start_dt"].dt.date.astype(str)
        df["appt_coord_str"] = df["appointment_id"].map(lambda x: appt_coord_map.get(str(x)))

        def compute_group(g: pd.DataFrame) -> pd.DataFrame:
            key = g.name if isinstance(g.name, tuple) else (g.name, None)
            tid = str(key[0]); date_key = key[1]
            g = g.sort_values("start_dt").copy()
            start_coord = office_coord_str if start_at_office else tech_home_map.get(tid, office_coord_str)
            prev_coords = g["appt_coord_str"].shift().to_numpy(dtype=object)
            if len(g) > 0:
                prev_coords[0] = start_coord
            curr_coords = g["appt_coord_str"].to_numpy(dtype=object)
            g["dist_from_prev_m"] = [
                _matrix_distance_m(matrix_distance, p, c)
                for p, c in zip(prev_coords, curr_coords)
            ]
            g["tech_id"] = tid
            g["date"] = date_key
            return g

        gb = df.groupby(["tech_id", "date"], group_keys=False, sort=True)
        try:
            df = gb.apply(compute_group, include_groups=False).reset_index(drop=True)
        except TypeError:
            df = gb.apply(compute_group).reset_index(drop=True)

        assigned_out = (df[[
            "appointment_id","status","start","end","tech_id","route_distance","route_duration","dist_from_prev_m"
        ]].sort_values(["tech_id","start"]))

    # ---------- Outliers: keep as single rows with blank tech_id ----------
    outliers_raw = raw.loc[mask_outlier, ["appointment_id","status","start","end","route_distance","route_duration"]].copy()
    if outliers_raw.empty:
        outliers_out = pd.DataFrame(columns=assigned_out.columns)
    else:
        outliers_out = outliers_raw.copy()
        outliers_out["tech_id"] = ""   # blank tech
        outliers_out["dist_from_prev_m"] = None
        outliers_out = outliers_out[assigned_out.columns] if not assigned_out.empty else outliers_out[[
            "appointment_id","status","start","end","tech_id","route_distance","route_duration","dist_from_prev_m"
        ]]

    # ---------- Combine ----------
    df_out = pd.concat([assigned_out, outliers_out], ignore_index=True)
    if not df_out.empty:
        df_out = df_out.sort_values(by=["tech_id","start"], na_position="last")

    # ---------- Add the 4 new columns from meta ----------
    def _from_meta(idx, i):
        meta = appt_meta.get(str(idx))
        return meta[i] if meta else None

    df_out["Çağrı tipi"]        = df_out["appointment_id"].map(lambda x: _from_meta(x, 0))
    df_out["Ürün Grubu"]        = df_out["appointment_id"].map(lambda x: _from_meta(x, 1))
    df_out["Yetkinlik Grubu"]   = df_out["appointment_id"].map(lambda x: _from_meta(x, 2))
    df_out["Randevu Tarih Saat"]= df_out["appointment_id"].map(lambda x: _from_meta(x, 3))

    # ---------- Insert placeholders for Excel formula columns (I & J) ----------
    df_out["cumulative distance"] = None
    df_out["total distance"] = None

    # Final order: keep E=tech_id, H=dist_from_prev_m, I/J for formula columns
    cols_final = [
        "appointment_id","status","start","end",
        "tech_id","route_distance","route_duration","dist_from_prev_m",
        "cumulative distance","total distance",
        "Çağrı tipi","Ürün Grubu","Yetkinlik Grubu","Randevu Tarih Saat",
    ]
    df_out = df_out.reindex(columns=cols_final)

    # Totals only for real technicians
    if assigned_out.empty:
        tech_totals = pd.DataFrame(columns=["tech_id","total_distance_m"])
    else:
        tech_totals = (assigned_out.groupby("tech_id")["dist_from_prev_m"]
                       .sum()
                       .reset_index()
                       .rename(columns={"dist_from_prev_m":"total_distance_m"}))

    return df_out, tech_totals

# ----------------- MAIN -----------------
def main():
    # Load JSONs
    result_obj = json.loads(Path(INPUT_RESULT_JSON).read_text(encoding="utf-8"))
    dl_obj     = json.loads(Path(INPUT_DATALOADER).read_text(encoding="utf-8"))

    assignments_df = flatten_assignments(result_obj.get("assignments"))
    # add any explicit outliers if provided separately
    outliers_df    = flatten_outliers(result_obj)
    if not outliers_df.empty:
        assignments_df = pd.concat([assignments_df, outliers_df], ignore_index=True)

    nonavail_df    = flatten_nonavail(result_obj.get("nonavailibilities"))
    techs_df       = flatten_technicians(result_obj.get("technicians"))
    sugg_dfs       = flatten_suggestions(result_obj.get("suggestions"))

    # NEW: assignments exploded per tech with hop distances (from MATRIX) + totals
    assignments_with_dist_df, tech_total_distance_df = build_assignments_with_dist(assignments_df, dl_obj)

    # Meta
    meta_df = pd.DataFrame({
        "assignments_count": [len(assignments_df)],
        "assignments_rows_on_sheet": [len(assignments_with_dist_df)],
        "technicians_count": [techs_df["technician_id"].nunique() if not techs_df.empty else 0],
        "nonavail_rows": [len(nonavail_df)],
    })

    # Tech load by slot (built from original assignments_df, outliers are ignored due to empty techs)
    tech_slot_counts_df    = build_tech_slot_counts(assignments_df)
    tech_slot_minutes_df   = build_tech_slot_minutes(assignments_df)

    # Maps for appt_eligibility_by_slot
    tech_used_map    = tech_slot_used_minutes_map(assignments_df)
    appt_overlap_map = appt_assigned_overlap_by_tech(assignments_df)
    appts            = dl_obj.get("appointments", [])
    appt_elig_df     = build_appt_eligibility_by_slot(appts, tech_used_map, appt_overlap_map)

    # Write all sheets in ONE file
    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        # Assignments sheet now per-tech + distance from previous hop (includes outliers)
        assignments_with_dist_df.to_excel(writer, index=False, sheet_name="Assignments")

        # --- NEW: add Excel formulas for cumulative & total distance ---
        wsA = writer.sheets["Assignments"]
        wsA.cell(row=1, column=9).value  = "cumulative distance"  # I1
        wsA.cell(row=1, column=10).value = "total distance"       # J1

        n_rows = len(assignments_with_dist_df) + 1  # + header
        if n_rows >= 2:
            for r in range(2, n_rows + 1):
                # I{r} = cumulative distance per tech_id
                wsA.cell(row=r, column=9).value  = f"=IF(E{r}=E{r-1},H{r}+I{r-1},H{r})"
                # J{r} = total distance written only on the last row of each tech_id block
                wsA.cell(row=r, column=10).value = f"=IF(E{r}=E{r+1},0,I{r})"

        nonavail_df.to_excel(writer, index=False, sheet_name="Nonavailabilities")
        techs_df.to_excel(writer, index=False, sheet_name="Technicians")
        meta_df.to_excel(writer, index=False, sheet_name="Meta")

        # Suggestions
        for name, df in sugg_dfs.items():
            df.to_excel(writer, index=False, sheet_name=name[:31])

        # Tech_Load_By_Slot (two tables one below another)
        sheet = "Tech_Load_By_Slot"
        start_row = 0
        pd.DataFrame({"Tech Load (Counts by Slot)": [""]}).to_excel(
            writer, index=False, header=False, sheet_name=sheet, startrow=start_row
        )
        start_row += 2
        tech_slot_counts_df.to_excel(writer, index=False, sheet_name=sheet, startrow=start_row)
        start_row += len(tech_slot_counts_df) + 3
        pd.DataFrame({"Tech Load (Duration minutes by Slot)": [""]}).to_excel(
            writer, index=False, header=False, sheet_name=sheet, startrow=start_row
        )
        start_row += 2
        tech_slot_minutes_df.to_excel(writer, index=False, sheet_name=sheet, startrow=start_row)

        # appt_eligibility_by_slot
        out_df = appt_elig_df.drop(columns=["is_arrival_window_row"])
        sheet2 = "appt_eligibility_by_slot"
        out_df.to_excel(writer, index=False, sheet_name=sheet2)

        # Bold arrival window rows
        try:
            wb = writer.book
            ws = writer.sheets[sheet2]
            bold_font = Font(bold=True)
            mask = appt_elig_df["is_arrival_window_row"].tolist()
            for i, is_bold in enumerate(mask, start=2):  # +1 header, +1 1-based
                if is_bold:
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=i, column=col).font = bold_font
        except Exception:
            pass

        # NEW: total distance per technician
        tech_total_distance_df.to_excel(writer, index=False, sheet_name="Tech_Total_Distance")

        # --- NEW: Objective sheet ---
        objective = result_obj.get("objective", {})
        # --- NEW: Objective sheet ---
        objective = result_obj.get("objective", {})
        if objective:
            obj_items = [{"Metric": k, "Value": v} for k, v in objective.items()]
            obj_df = pd.DataFrame(obj_items)
            obj_df.to_excel(writer, index=False, sheet_name="Objective", header=False, startrow=0, startcol=0)

            obj_df.to_excel(writer, index=False, sheet_name="Objective")

    print(f"✅ Wrote Excel with all sheets to: {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()

